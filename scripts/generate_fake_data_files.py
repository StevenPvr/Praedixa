"""Generate synthetic HR/logistics data for 3 very different clients.

Each client has a unique data culture, file structure, naming conventions,
column names, date formats, separators, and level of data quality.

Clients:
  1. Acme Logistics — Large 3PL, professional tools (Lucca + PayFit)
     Clean exports, semicolons, DD/MM/YYYY, UTF-8-BOM, PascalCase Lucca headers
  2. TransFroid Express — Medium cold-chain transport, homemade Excel
     Messy spreadsheets, tabs/spaces in names, mixed date formats, accents
  3. Petit Colis SARL — Small e-commerce fulfillment, Google Sheets
     Comma separator, ISO dates mixed with French, missing data, English headers

Output: data/{client-slug}/...

Usage:
    cd apps/api
    uv run python ../../scripts/generate_fake_data_files.py
"""

import csv
import json
import random
from datetime import date, timedelta
from pathlib import Path

from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

fake = Faker("fr_FR")
Faker.seed(42)
random.seed(42)

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"

# ════════════════════════════════════════════════════════
#  Shared helpers
# ════════════════════════════════════════════════════════

def fr_date(d: date) -> str:
    return d.strftime("%d/%m/%Y")

def iso_date(d: date) -> str:
    return d.isoformat()

def fr_decimal(v: float) -> str:
    return f"{v:.2f}".replace(".", ",")

def en_decimal(v: float) -> str:
    return f"{v:.2f}"

def write_csv(
    path: Path, rows: list[dict], *,
    delimiter: str = ";", encoding: str = "utf-8-sig",
) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        return 0
    fields = list(rows[0].keys())
    with open(path, "w", encoding=encoding, newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter=delimiter)
        w.writeheader()
        w.writerows(rows)
    return len(rows)

def write_xlsx(path: Path, sheets: dict[str, list[dict]], bold_header: bool = False) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    total = 0
    for i, (sheet_name, rows) in enumerate(sheets.items()):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = sheet_name
        if not rows:
            continue
        fields = list(rows[0].keys())
        ws.append(fields)
        if bold_header:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
                cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([row.get(f, "") for f in fields])
        total += len(rows)
    wb.save(path)
    return total

def seasonal_absence_factor(d: date) -> float:
    if d.month in (7, 8):
        return 2.5
    if d.month == 12 and d.day >= 20:
        return 0.4
    return 1.0

def volume_factor(d: date) -> float:
    m = d.month
    if m == 12: return 1.6
    if m in (7, 8): return 0.65
    if m == 11: return 1.3
    if m in (3, 9): return 1.15
    return 1.0


# ════════════════════════════════════════════════════════
#  CLIENT 1 — Acme Logistics (large 3PL, professional)
# ════════════════════════════════════════════════════════
#
#  Structure:
#    acme-logistics/
#    ├── sirh/
#    │   ├── effectifs_2025.csv          (;  DD/MM/YYYY  UTF-8-BOM)
#    │   ├── absences_2024_2025.csv
#    │   └── planning_s47_s48.csv
#    ├── lucca/
#    │   ├── Lucca_Collaborateurs_Export.csv
#    │   ├── Lucca_Conges_2025.csv
#    │   └── Lucca_Compteurs_Soldes.csv
#    ├── payfit/
#    │   ├── PayFit_Bulletins_2025.csv
#    │   └── PayFit_Absences_2025.csv
#    ├── operations/
#    │   ├── volumes_journaliers.csv
#    │   └── meteo_sites.csv
#    └── finance/
#        ├── factures_clients_2025.csv
#        └── budget_2025.csv

ACME_SITES = [
    {"name": "Entrepot Lyon", "code": "LYO", "city": "Lyon"},
    {"name": "Entrepot Marseille", "code": "MRS", "city": "Marseille"},
    {"name": "Entrepot Paris-Nord", "code": "CDG", "city": "Roissy"},
]
ACME_DEPTS = [
    "Reception", "Preparation", "Expedition",
    "Controle Qualite", "Maintenance", "Transport", "Administration",
]
ACME_POSTES = [
    "Cariste", "Preparateur de commandes", "Agent de quai",
    "Chef d'equipe", "Responsable logistique", "Technicien maintenance",
    "Agent administratif", "Chauffeur PL", "Controleur qualite",
]
ACME_ABSENCE_TYPES = [
    ("CP", "Conge paye"), ("RTT", "RTT"), ("MAL", "Arret maladie"),
    ("AT", "Accident du travail"), ("MAT", "Conge maternite"),
    ("PAT", "Conge paternite"), ("CSS", "Conge sans solde"),
    ("FOR", "Formation"), ("ABS", "Absence injustifiee"),
]

def gen_acme():
    out = DATA / "acme-logistics"
    stats = []

    # ── Effectifs ───────────────────────────────
    employees = []
    for i in range(350):
        site = random.choice(ACME_SITES)
        employees.append({
            "Matricule": f"ACM{i + 1:04d}",
            "Nom": fake.last_name(),
            "Prenom": fake.first_name(),
            "Date de naissance": fr_date(fake.date_between("-55y", "-20y")),
            "Date d'embauche": fr_date(fake.date_between("-10y", "-30d")),
            "Site": site["name"],
            "Code site": site["code"],
            "Departement": random.choice(ACME_DEPTS),
            "Poste": random.choice(ACME_POSTES),
            "Type contrat": random.choices(
                ["CDI", "CDD", "Interim", "Alternance"],
                weights=[65, 15, 15, 5], k=1,
            )[0],
            "Temps de travail": fr_decimal(random.choice([1.0, 1.0, 1.0, 0.8, 0.5])),
            "Email": f"acm{i + 1:04d}@acme-logistics.fr",
            "Telephone": fake.phone_number(),
        })
    stats.append(("sirh/effectifs_2025.csv",
                   write_csv(out / "sirh" / "effectifs_2025.csv", employees)))

    # ── Absences ────────────────────────────────
    absences = []
    d = date(2024, 1, 1)
    while d <= date(2025, 12, 31):
        n = max(0, int(random.gauss(3, 1.5) * seasonal_absence_factor(d)))
        for _ in range(n):
            emp = random.choice(employees)
            code, label = random.choice(ACME_ABSENCE_TYPES)
            dur = random.choices([1, 2, 3, 5, 10, 20], weights=[30, 20, 15, 20, 10, 5], k=1)[0]
            absences.append({
                "Matricule": emp["Matricule"],
                "Nom": emp["Nom"],
                "Prenom": emp["Prenom"],
                "Site": emp["Site"],
                "Departement": emp["Departement"],
                "Date debut": fr_date(d),
                "Date fin": fr_date(d + timedelta(days=dur - 1)),
                "Nb jours": str(dur),
                "Code absence": code,
                "Libelle": label,
                "Motif": "Certificat medical" if code in ("MAL", "AT", "MAT", "PAT") else "",
                "Statut": random.choices(
                    ["Valide", "En attente", "Refuse"], weights=[80, 15, 5], k=1
                )[0],
            })
        d += timedelta(days=1)
    stats.append(("sirh/absences_2024_2025.csv",
                   write_csv(out / "sirh" / "absences_2024_2025.csv", absences)))

    # ── Planning ────────────────────────────────
    shifts = [("06:00", "14:00"), ("14:00", "22:00"), ("22:00", "06:00"),
              ("08:00", "16:00"), ("09:00", "17:00")]
    planning = []
    for off in range(14):
        day = date(2025, 11, 17) + timedelta(days=off)
        if day.weekday() == 6:
            continue
        for emp in employees:
            if random.random() < 0.12:
                continue
            s = random.choices(shifts, weights=[25, 25, 10, 20, 20], k=1)[0]
            planning.append({
                "Matricule": emp["Matricule"],
                "Nom complet": f"{emp['Prenom']} {emp['Nom']}",
                "Date": fr_date(day),
                "Heure debut": s[0],
                "Heure fin": s[1],
                "Site": emp["Site"],
                "Departement": emp["Departement"],
                "Poste": emp["Poste"],
            })
    stats.append(("sirh/planning_s47_s48.csv",
                   write_csv(out / "sirh" / "planning_s47_s48.csv", planning)))

    # ── Lucca exports ───────────────────────────
    lucca_collab = [
        {"Id": str(random.randint(10000, 99999)),
         "Nom": e["Nom"], "Prenom": e["Prenom"], "Email": e["Email"],
         "DateEmbauche": e["Date d'embauche"], "Departement": e["Departement"],
         "Etablissement": e["Site"], "Statut": "Actif",
         "Manager": fake.name(), "TypeContrat": e["Type contrat"]}
        for e in employees
    ]
    stats.append(("lucca/Lucca_Collaborateurs_Export.csv",
                   write_csv(out / "lucca" / "Lucca_Collaborateurs_Export.csv", lucca_collab)))

    lucca_conges = [
        {"Id": str(random.randint(100000, 999999)),
         "Collaborateur": f"{a['Prenom']} {a['Nom']}",
         "DateDebut": a["Date debut"], "DateFin": a["Date fin"],
         "NbJours": a["Nb jours"],
         "Type": a["Libelle"],
         "Statut": {"Valide": "Approuve", "En attente": "En attente",
                    "Refuse": "Refuse"}[a["Statut"]],
         "ValidePar": fake.name() if a["Statut"] == "Valide" else ""}
        for a in absences[:2500]
    ]
    stats.append(("lucca/Lucca_Conges_2025.csv",
                   write_csv(out / "lucca" / "Lucca_Conges_2025.csv", lucca_conges)))

    lucca_compteurs = [
        {"Collaborateur": f"{e['Prenom']} {e['Nom']}",
         "TypeCompteur": ct, "Acquis": fr_decimal(acq),
         "Pris": fr_decimal(pris), "Solde": fr_decimal(acq - pris),
         "DateMaj": fr_date(date(2025, 12, 31))}
        for e in employees
        for ct, acq, pris in [
            ("Conges payes", 25.0, round(random.uniform(5, 25), 1)),
            ("RTT", 10.0, round(random.uniform(0, 10), 1)),
        ]
    ]
    stats.append(("lucca/Lucca_Compteurs_Soldes.csv",
                   write_csv(out / "lucca" / "Lucca_Compteurs_Soldes.csv", lucca_compteurs)))

    # ── PayFit exports ──────────────────────────
    payfit_bul = []
    for month in range(1, 13):
        for emp in random.sample(employees, 200):
            brut = round(random.uniform(1800, 4500), 2)
            charges = round(brut * random.uniform(0.20, 0.25), 2)
            payfit_bul.append({
                "Mois": f"01/{month:02d}/2025",
                "Matricule": emp["Matricule"], "Nom": emp["Nom"],
                "Prenom": emp["Prenom"],
                "BrutMensuel": fr_decimal(brut),
                "CotisationsSalariales": fr_decimal(charges),
                "NetAPayer": fr_decimal(brut - charges),
                "HeuresTravaillees": fr_decimal(
                    float(emp["Temps de travail"].replace(",", ".")) * 151.67),
                "CongesRestants": str(random.randint(0, 25)),
            })
    stats.append(("payfit/PayFit_Bulletins_2025.csv",
                   write_csv(out / "payfit" / "PayFit_Bulletins_2025.csv", payfit_bul)))

    payfit_abs = [
        {"Matricule": a["Matricule"],
         "NomComplet": f"{a['Nom']} {a['Prenom']}",
         "Debut": a["Date debut"], "Fin": a["Date fin"],
         "Duree": a["Nb jours"], "Nature": a["Libelle"],
         "ImpactPaie": "Oui" if a["Code absence"] in ("MAL", "AT", "CSS", "ABS") else "Non"}
        for a in absences
    ]
    stats.append(("payfit/PayFit_Absences_2025.csv",
                   write_csv(out / "payfit" / "PayFit_Absences_2025.csv", payfit_abs)))

    # ── Volumes ─────────────────────────────────
    volumes = []
    d = date(2024, 1, 1)
    while d <= date(2025, 12, 31):
        if d.weekday() < 6:
            wkd_f = 0.4 if d.weekday() == 5 else 1.0
            for site in ACME_SITES:
                for dept in ["Reception", "Preparation", "Expedition"]:
                    base = {"Reception": 820, "Preparation": 780, "Expedition": 710}[dept]
                    nb = max(0, int(random.gauss(base, 100) * volume_factor(d) * wkd_f))
                    volumes.append({
                        "Date": fr_date(d), "Site": site["name"],
                        "Code site": site["code"], "Departement": dept,
                        "Nb colis": str(nb),
                        "Nb palettes": str(nb // random.randint(15, 25)),
                        "Poids total (kg)": fr_decimal(nb * random.uniform(1.5, 4.5)),
                        "Nb erreurs": str(random.randint(0, 5)),
                        "Taux de service (%)": fr_decimal(random.uniform(95, 99.9)),
                    })
        d += timedelta(days=1)
    stats.append(("operations/volumes_journaliers.csv",
                   write_csv(out / "operations" / "volumes_journaliers.csv", volumes)))

    # ── Meteo ───────────────────────────────────
    meteo = []
    d = date(2024, 1, 1)
    base_temp = {1: 3, 2: 4, 3: 8, 4: 12, 5: 16, 6: 20,
                 7: 23, 8: 22, 9: 18, 10: 13, 11: 7, 12: 4}
    city_off = {"Lyon": 0, "Marseille": 3, "Roissy": -1}
    while d <= date(2025, 12, 31):
        for site in ACME_SITES:
            t = base_temp[d.month] + city_off[site["city"]] + random.gauss(0, 2)
            meteo.append({
                "Date": fr_date(d), "Site": site["name"],
                "Temperature min (C)": fr_decimal(t - random.uniform(3, 7)),
                "Temperature max (C)": fr_decimal(t + random.uniform(3, 7)),
                "Temperature moy (C)": fr_decimal(t),
                "Humidite (%)": str(random.randint(40, 95)),
                "Precipitation (mm)": fr_decimal(max(0, random.gauss(2, 3))),
            })
        d += timedelta(days=1)
    stats.append(("operations/meteo_sites.csv",
                   write_csv(out / "operations" / "meteo_sites.csv", meteo)))

    # ── Finance ─────────────────────────────────
    factures = []
    for month in range(1, 13):
        for _ in range(random.randint(20, 40)):
            ht = round(random.uniform(500, 15000), 2)
            tva = round(ht * 0.20, 2)
            factures.append({
                "N° facture": f"FC-2025-{len(factures) + 1:05d}",
                "Date emission": fr_date(date(2025, month, random.randint(1, 28))),
                "Client": fake.company(),
                "Description": random.choice([
                    "Prestation logistique", "Transport express",
                    "Stockage mensuel", "Preparation commandes",
                ]),
                "Montant HT": fr_decimal(ht),
                "TVA 20%": fr_decimal(tva),
                "Montant TTC": fr_decimal(ht + tva),
                "Statut": random.choices(
                    ["Payee", "En attente", "En retard"], weights=[70, 20, 10], k=1
                )[0],
            })
    stats.append(("finance/factures_clients_2025.csv",
                   write_csv(out / "finance" / "factures_clients_2025.csv", factures)))

    budget = []
    cats = [("Masse salariale", 150000, 180000), ("Transport", 30000, 50000),
            ("Maintenance", 10000, 20000), ("Energie", 8000, 15000),
            ("Fournitures", 3000, 8000), ("Sous-traitance", 20000, 40000)]
    for month in range(1, 13):
        for site in ACME_SITES:
            for cat, lo, hi in cats:
                bud = round(random.uniform(lo, hi) / 12, 2)
                reel = round(bud * random.uniform(0.85, 1.15), 2)
                budget.append({
                    "Mois": f"01/{month:02d}/2025", "Site": site["name"],
                    "Categorie": cat, "Budget prevu": fr_decimal(bud),
                    "Reel": fr_decimal(reel), "Ecart": fr_decimal(reel - bud),
                    "Ecart %": fr_decimal((reel - bud) / bud * 100),
                })
    stats.append(("finance/budget_2025.csv",
                   write_csv(out / "finance" / "budget_2025.csv", budget)))

    return stats


# ════════════════════════════════════════════════════════
#  CLIENT 2 — TransFroid Express (medium, messy Excel)
# ════════════════════════════════════════════════════════
#
#  Structure:
#    transfroid-express/
#    ├── RH - Données Personnel/
#    │   └── Fichier_RH_Master_2025.xlsx      (multi-sheet!)
#    ├── Absences et Congés/
#    │   ├── suivi_absences_2024.xlsx
#    │   └── suivi_absences_2025.xlsx
#    ├── Planning Tournées/
#    │   └── Planning_Chauffeurs_Dec2025.xlsx
#    ├── Exploitation/
#    │   ├── tonnages_quotidiens.csv           (tab-separated!)
#    │   └── incidents_frigo.csv               (tab-separated!)
#    └── Données Températures/
#        └── releves_temperature_camions.csv   (tab-separated!)
#
#  Quirks: tab separator, DD-MM-YYYY dates, column names with
#  accents and spaces, multi-sheet Excel, merged cells style,
#  some empty rows, inconsistent casing

TF_SITES = [
    {"name": "Plateforme Rungis", "code": "RGS"},
    {"name": "Dépôt Bordeaux", "code": "BDX"},
]
TF_SERVICES = ["Frais", "Surgelé", "Sec", "Ultra-frais"]
TF_POSTES_CHAUFFEUR = [
    "Chauffeur PL Frigo", "Chauffeur SPL", "Livreur VL",
    "Chef de quai", "Préparateur froid", "Cariste frigo",
    "Agent de quai", "Responsable exploitation",
]

def gen_transfroid():
    out = DATA / "transfroid-express"
    stats = []

    # ── RH Master (multi-sheet XLSX) ────────────
    employees = []
    for i in range(130):
        site = random.choice(TF_SITES)
        employees.append({
            "N° Salarié": f"TF-{i + 1:03d}",
            "NOM": fake.last_name().upper(),  # They use UPPERCASE names
            "Prénom": fake.first_name(),
            "Date Naissance": fake.date_between("-55y", "-20y").strftime("%d-%m-%Y"),
            "Entrée Société": fake.date_between("-12y", "-60d").strftime("%d-%m-%Y"),
            "Site Rattachement": site["name"],
            "Service": random.choice(TF_SERVICES),
            "Fonction": random.choice(TF_POSTES_CHAUFFEUR),
            "Contrat": random.choices(
                ["CDI", "CDD 6 mois", "CDD 12 mois", "Intérim"],
                weights=[60, 15, 10, 15], k=1
            )[0],
            "Tps Travail": random.choice(["100%", "100%", "100%", "80%", "50%"]),
            "Permis": random.choice(["B", "C", "CE", "C+FIMO", "CE+FIMO", "B"]),
            "FIMO/FCO valide": random.choice(["oui", "non", "oui", "oui", "N/A"]),
            "Tél portable": fake.phone_number(),
            "Observations": random.choice(["", "", "", "Habilitation CACES 3",
                                           "Restriction port de charges", ""]),
        })

    # Second sheet: coordonnées
    coordonnees = [
        {"N° Salarié": e["N° Salarié"],
         "NOM Prénom": f"{e['NOM']} {e['Prénom']}",
         "Adresse": fake.street_address(),
         "CP": fake.postcode(),
         "Ville": fake.city(),
         "Email perso": fake.free_email(),
         "Contact urgence": fake.name(),
         "Tél urgence": fake.phone_number()}
        for e in employees
    ]

    # Third sheet: véhicules assignés
    vehicules = []
    for e in employees:
        if "Chauffeur" in e["Fonction"] or "Livreur" in e["Fonction"]:
            vehicules.append({
                "N° Salarié": e["N° Salarié"],
                "Conducteur": f"{e['NOM']} {e['Prénom']}",
                "Immatriculation": fake.license_plate(),
                "Type véhicule": random.choice([
                    "Renault Master Frigo", "Iveco Daily -20°C",
                    "Scania R450 Frigo", "DAF XF Bi-temp",
                    "Mercedes Sprinter Frigo",
                ]),
                "PTAC (t)": random.choice(["3,5", "7,5", "19", "26", "44"]),
                "Date CT": fake.date_between("-1y", "+6m").strftime("%d-%m-%Y"),
            })

    n = write_xlsx(out / "RH - Données Personnel" / "Fichier_RH_Master_2025.xlsx",
                   {"Personnel": employees, "Coordonnées": coordonnees,
                    "Véhicules": vehicules}, bold_header=True)
    stats.append(("RH - Données Personnel/Fichier_RH_Master_2025.xlsx", n))

    # ── Absences (yearly XLSX, messy) ───────────
    for year in [2024, 2025]:
        abs_rows = []
        d = date(year, 1, 1)
        end = date(year, 12, 31)
        while d <= end:
            n_abs = max(0, int(random.gauss(1.5, 1) * seasonal_absence_factor(d)))
            for _ in range(n_abs):
                emp = random.choice(employees)
                dur = random.choices([0.5, 1, 2, 3, 5, 10], weights=[10, 30, 20, 15, 15, 10], k=1)[0]
                type_abs = random.choice([
                    "CP", "RTT", "Maladie", "Maladie pro", "AT",
                    "Maternité", "Paternité", "Congé sans solde",
                    "Récup heures", "Événement familial",
                ])
                abs_rows.append({
                    "N° Salarié": emp["N° Salarié"],
                    "NOM": emp["NOM"],
                    "Prénom": emp["Prénom"],
                    "Site": emp["Site Rattachement"],
                    "Du": d.strftime("%d-%m-%Y"),
                    "Au": (d + timedelta(days=max(0, int(dur) - 1))).strftime("%d-%m-%Y"),
                    "Nb jrs": str(dur).replace(".", ","),
                    "Type": type_abs,
                    "Justificatif": random.choice(["Oui", "Non", "Oui", "Oui", "En attente"]),
                    "Saisi par": random.choice(["M. DUPONT", "Mme MARTIN", "RH", "Auto"]),
                    "Remarque": random.choice(["", "", "", "prolongation",
                                                "mi-temps thérapeutique", ""]),
                })
            d += timedelta(days=1)
        n = write_xlsx(out / "Absences et Congés" / f"suivi_absences_{year}.xlsx",
                       {f"Absences {year}": abs_rows}, bold_header=True)
        stats.append((f"Absences et Congés/suivi_absences_{year}.xlsx", n))

    # ── Planning Chauffeurs (XLSX) ──────────────
    plan_rows = []
    for off in range(21):  # 3 weeks
        day = date(2025, 12, 1) + timedelta(days=off)
        if day.weekday() == 6:
            continue
        for emp in employees:
            if "Chauffeur" not in emp["Fonction"] and "Livreur" not in emp["Fonction"]:
                continue
            if random.random() < 0.15:
                plan_rows.append({
                    "Date": day.strftime("%d-%m-%Y"),
                    "N° Salarié": emp["N° Salarié"],
                    "Conducteur": f"{emp['NOM']} {emp['Prénom']}",
                    "Tournée": "REPOS" if day.weekday() == 5 else "ABSENT",
                    "Départ": "", "Retour prévu": "",
                    "Km estimés": "", "Nb points livraison": "",
                })
                continue
            depart = random.choice(["03:00", "04:00", "05:00", "06:00", "14:00"])
            km = random.randint(80, 450)
            pts = random.randint(5, 25)
            plan_rows.append({
                "Date": day.strftime("%d-%m-%Y"),
                "N° Salarié": emp["N° Salarié"],
                "Conducteur": f"{emp['NOM']} {emp['Prénom']}",
                "Tournée": random.choice(["T1-RGS", "T2-RGS", "T3-BDX", "T4-BDX",
                                          "INTER-RGS-BDX", "LIVR-IDF", "LIVR-PACA"]),
                "Départ": depart,
                "Retour prévu": f"{int(depart[:2]) + random.randint(8, 12):02d}:00",
                "Km estimés": str(km),
                "Nb points livraison": str(pts),
            })
    n = write_xlsx(out / "Planning Tournées" / "Planning_Chauffeurs_Dec2025.xlsx",
                   {"Planning Dec 2025": plan_rows}, bold_header=True)
    stats.append(("Planning Tournées/Planning_Chauffeurs_Dec2025.xlsx", n))

    # ── Tonnages (tab-separated CSV!) ───────────
    tonnages = []
    d = date(2024, 1, 1)
    while d <= date(2025, 12, 31):
        if d.weekday() < 6:
            for site in TF_SITES:
                for service in TF_SERVICES:
                    base = {"Frais": 12, "Surgelé": 8, "Sec": 15, "Ultra-frais": 5}[service]
                    t = max(0, round(random.gauss(base, 3) * volume_factor(d), 1))
                    tonnages.append({
                        "date": d.strftime("%d-%m-%Y"),
                        "site": site["name"],
                        "service": service,
                        "tonnage (t)": str(t).replace(".", ","),
                        "nb tournées": str(random.randint(3, 15)),
                        "nb clients livrés": str(random.randint(20, 80)),
                        "nb litiges": str(random.randint(0, 3)),
                        "rupture chaîne froid": random.choice(["non", "non", "non", "non", "oui"]),
                    })
        d += timedelta(days=1)
    stats.append(("Exploitation/tonnages_quotidiens.csv",
                   write_csv(out / "Exploitation" / "tonnages_quotidiens.csv",
                             tonnages, delimiter="\t", encoding="utf-8")))

    # ── Incidents frigo (tab-separated) ─────────
    incidents = []
    for _ in range(85):
        d = fake.date_between(date(2024, 1, 1), date(2025, 12, 31))
        incidents.append({
            "date incident": d.strftime("%d-%m-%Y"),
            "site": random.choice(TF_SITES)["name"],
            "zone": random.choice(["Quai réception", "Chambre froide 1",
                                   "Chambre froide 2", "Camion", "Quai expédition"]),
            "température relevée (°C)": str(
                round(random.uniform(-25, 8), 1)).replace(".", ","),
            "seuil dépassé (°C)": random.choice(["-18", "-18", "2", "4", "7"]),
            "durée dépassement (min)": str(random.randint(5, 120)),
            "action corrective": random.choice([
                "Transfert produits", "Réparation groupe froid",
                "Destruction lot", "Reconditionnement",
                "Alerte fournisseur",
            ]),
            "gravité": random.choice(["Mineure", "Mineure", "Majeure", "Critique"]),
        })
    stats.append(("Exploitation/incidents_frigo.csv",
                   write_csv(out / "Exploitation" / "incidents_frigo.csv",
                             incidents, delimiter="\t", encoding="utf-8")))

    # ── Températures camions (tab-separated) ────
    temp_rows = []
    d = date(2025, 1, 1)
    while d <= date(2025, 12, 31):
        if d.weekday() < 6:
            for v in vehicules[:30]:
                for zone in ["Compartiment 1", "Compartiment 2"]:
                    target = random.choice([-20, -18, 2, 4])
                    temp_rows.append({
                        "date": d.strftime("%d-%m-%Y"),
                        "immatriculation": v["Immatriculation"],
                        "type véhicule": v["Type véhicule"],
                        "zone": zone,
                        "consigne (°C)": str(target),
                        "t° min relevée": str(
                            round(target + random.uniform(-2, 0.5), 1)).replace(".", ","),
                        "t° max relevée": str(
                            round(target + random.uniform(0, 3), 1)).replace(".", ","),
                        "conforme": random.choices(
                            ["OUI", "NON"], weights=[92, 8], k=1)[0],
                    })
        d += timedelta(days=1)
    stats.append(("Données Températures/releves_temperature_camions.csv",
                   write_csv(out / "Données Températures" / "releves_temperature_camions.csv",
                             temp_rows, delimiter="\t", encoding="utf-8")))

    return stats


# ════════════════════════════════════════════════════════
#  CLIENT 3 — Petit Colis SARL (small, Google Sheets)
# ════════════════════════════════════════════════════════
#
#  Structure:
#    petit-colis/
#    ├── team/
#    │   └── employees.csv              (comma-sep, English headers!)
#    ├── hr/
#    │   ├── absences_export.csv        (comma, ISO dates sometimes)
#    │   └── leave_balances.csv         (comma, English)
#    ├── ops/
#    │   ├── daily_shipments.csv        (comma, English)
#    │   └── returns_log.csv            (comma, messy)
#    └── misc/
#        └── notes_reunions.csv         (French free text, commas)
#
#  Quirks: comma separator (Google Sheets export), English column names,
#  ISO dates for some files / French for others, missing values,
#  trailing spaces, inconsistent capitalization, some rows have
#  extra commas from free-text fields

PC_ROLES = ["Picker", "Packer", "Shipping agent", "Team lead",
            "Warehouse manager", "Returns specialist", "IT support"]

def gen_petitcolis():
    out = DATA / "petit-colis"
    stats = []

    # ── Employees (English headers, comma) ──────
    employees = []
    for i in range(48):
        hire = fake.date_between("-5y", "-30d")
        employees.append({
            "employee_id": f"PC{i + 1:03d}",
            "first_name": fake.first_name(),
            "last_name": fake.last_name(),
            "email": f"pc{i + 1:03d}@petitcolis.fr",
            "hire_date": iso_date(hire),
            "role": random.choice(PC_ROLES),
            "contract": random.choices(
                ["permanent", "fixed-term", "intern"], weights=[70, 20, 10], k=1
            )[0],
            "hours_per_week": str(random.choice([35, 35, 35, 28, 20])),
            "manager": random.choice(["Sophie L.", "Marc D.", "Julie P."]),
            # Typical Google Sheets messiness:
            "notes": random.choice([
                "", "", "", "  ", "started as intern",
                "bilingual EN/FR", "forklift certified",
                "on probation until " + iso_date(hire + timedelta(days=90)),
            ]),
        })
    stats.append(("team/employees.csv",
                   write_csv(out / "team" / "employees.csv", employees,
                             delimiter=",", encoding="utf-8")))

    # ── Absences (mixed date formats!) ──────────
    absences = []
    d = date(2025, 1, 1)
    while d <= date(2025, 12, 31):
        n = max(0, int(random.gauss(0.5, 0.5) * seasonal_absence_factor(d)))
        for _ in range(n):
            emp = random.choice(employees)
            dur = random.choices([0.5, 1, 2, 3, 5], weights=[15, 35, 25, 15, 10], k=1)[0]
            end_d = d + timedelta(days=max(0, int(dur) - 1))
            # Mix of ISO and French dates (realistic messiness)
            use_iso = random.random() < 0.6
            absences.append({
                "employee_id": emp["employee_id"],
                "name": f"{emp['first_name']} {emp['last_name']}",
                "start_date": iso_date(d) if use_iso else fr_date(d),
                "end_date": iso_date(end_d) if use_iso else fr_date(end_d),
                "days": str(dur),
                "type": random.choice([
                    "paid_leave", "sick", "unpaid", "family_event",
                    "RTT", "remote_work",  # remote_work shouldn't be here but it's realistic
                    "maternity", "paternity",
                ]),
                "status": random.choice(["approved", "approved", "approved",
                                          "pending", "rejected"]),
                "approved_by": random.choice(["Sophie L.", "Marc D.", ""]),
                # Some rows have notes with commas (will cause issues!)
                "comment": random.choice([
                    "", "", "", "", "doctor's note pending",
                    "approved, but check dates",  # comma in value!
                    "half day AM", "half day PM",
                    "COVID, PCR positive",  # comma in value!
                ]),
            })
        d += timedelta(days=1)
    stats.append(("hr/absences_export.csv",
                   write_csv(out / "hr" / "absences_export.csv", absences,
                             delimiter=",", encoding="utf-8")))

    # ── Leave balances (English) ────────────────
    balances = [
        {"employee_id": e["employee_id"],
         "name": f"{e['first_name']} {e['last_name']}",
         "paid_leave_total": str(25),
         "paid_leave_taken": str(random.randint(5, 25)),
         "paid_leave_remaining": "",  # intentionally blank, should be computed
         "rtt_total": str(10 if e["contract"] == "permanent" else 0),
         "rtt_taken": str(random.randint(0, 10)),
         "sick_days_ytd": str(random.randint(0, 12)),
         "last_updated": iso_date(date(2025, 12, 15))}
        for e in employees
    ]
    # Fix remaining (leave some blank as real data would)
    for b in balances:
        if random.random() < 0.8:
            b["paid_leave_remaining"] = str(
                int(b["paid_leave_total"]) - int(b["paid_leave_taken"]))
    stats.append(("hr/leave_balances.csv",
                   write_csv(out / "hr" / "leave_balances.csv", balances,
                             delimiter=",", encoding="utf-8")))

    # ── Daily shipments (English, ISO dates) ────
    shipments = []
    d = date(2025, 1, 1)
    while d <= date(2025, 12, 31):
        if d.weekday() < 5:  # Mon-Fri only
            nb = max(0, int(random.gauss(180, 40) * volume_factor(d)))
            shipments.append({
                "date": iso_date(d),
                "orders_received": str(nb + random.randint(-20, 20)),
                "orders_shipped": str(nb),
                "orders_returned": str(random.randint(2, 15)),
                "avg_weight_kg": en_decimal(random.uniform(0.3, 2.5)),
                "carriers": random.choice([
                    "Colissimo", "Chronopost", "Mondial Relay",
                    "Colissimo;Chronopost",  # multiple carriers, bad separator
                    "DPD", "GLS",
                ]),
                "picking_errors": str(random.randint(0, 4)),
                "packing_time_avg_min": en_decimal(random.uniform(1.5, 4.0)),
            })
        d += timedelta(days=1)
    stats.append(("ops/daily_shipments.csv",
                   write_csv(out / "ops" / "daily_shipments.csv", shipments,
                             delimiter=",", encoding="utf-8")))

    # ── Returns log (messy!) ────────────────────
    returns = []
    for _ in range(random.randint(200, 350)):
        d = fake.date_between(date(2025, 1, 1), date(2025, 12, 31))
        returns.append({
            "return_id": f"RET-{random.randint(10000, 99999)}",
            "date": iso_date(d) if random.random() < 0.7 else fr_date(d),
            "order_ref": f"ORD-{random.randint(100000, 999999)}",
            "reason": random.choice([
                "wrong item", "damaged", "too late", "changed mind",
                "defective", "missing parts", "wrong size",
                "never received",  # shouldn't be a return but happens
                "",  # missing reason
            ]),
            "refund_amount": en_decimal(random.uniform(5, 120)) if random.random() < 0.9 else "",
            "carrier": random.choice(["Colissimo", "Chronopost", "Mondial Relay",
                                       "client drop-off", ""]),
            "processed": random.choice(["yes", "no", "yes", "yes", "partial", ""]),
            "notes": random.choice([
                "", "", "customer angry", "sent replacement",
                "item resellable", "item destroyed",
                "see ticket #" + str(random.randint(1000, 9999)),
            ]),
        })
    stats.append(("ops/returns_log.csv",
                   write_csv(out / "ops" / "returns_log.csv", returns,
                             delimiter=",", encoding="utf-8")))

    # ── Notes réunions (French free text) ───────
    reunions = []
    for month in range(1, 13):
        for _ in range(random.randint(1, 3)):
            day = random.randint(1, 28)
            reunions.append({
                "date": fr_date(date(2025, month, day)),
                "type": random.choice(["Hebdo ops", "Point RH", "Comité direction",
                                        "Brief matin", "Rétrospective"]),
                "participants": "; ".join(random.sample(
                    ["Sophie L.", "Marc D.", "Julie P.", "Thomas R.",
                     "Émilie B.", "Nicolas G."], k=random.randint(2, 5))),
                "sujets": random.choice([
                    "Pic de Noël, recrutement 3 intérimaires",
                    "Retards Chronopost, basculer sur DPD ?",
                    "Nouvel outil WMS à tester en janvier",
                    "Bilan absences Q3, trop de RTT en même temps",
                    "Formation sécurité obligatoire à planifier",
                    "Problème picking zone B, réorg nécessaire",
                    "Budget 2026 à finaliser avant le 15",
                ]),
                "actions": random.choice([
                    "Marc contacte agence intérim",
                    "Sophie fait le point avec Chronopost",
                    "Julie planifie les formations",
                    "Thomas prépare le budget",
                    "RAS",
                    "",
                ]),
            })
    stats.append(("misc/notes_reunions.csv",
                   write_csv(out / "misc" / "notes_reunions.csv", reunions,
                             delimiter=",", encoding="utf-8")))

    return stats


# ════════════════════════════════════════════════════════
#  Metadata & Main
# ════════════════════════════════════════════════════════

def write_metadata():
    meta = {
        "acme-logistics": {
            "display_name": "Acme Logistics",
            "industry": "3PL / Entreposage",
            "employees": 350,
            "sites": 3,
            "data_culture": "Professional — Lucca + PayFit + SIRH structure",
            "file_conventions": "CSV semicolon, DD/MM/YYYY, UTF-8-BOM, French decimals",
        },
        "transfroid-express": {
            "display_name": "TransFroid Express",
            "industry": "Transport frigorifique",
            "employees": 130,
            "sites": 2,
            "data_culture": "Mixed — Excel multi-onglets, tab-separated CSV, homemade",
            "file_conventions": "XLSX multi-sheet, TSV, DD-MM-YYYY, accents in names",
        },
        "petit-colis": {
            "display_name": "Petit Colis SARL",
            "industry": "E-commerce fulfillment",
            "employees": 48,
            "sites": 1,
            "data_culture": "Basic — Google Sheets exports, English headers, messy",
            "file_conventions": "CSV comma, ISO/mixed dates, English, missing values",
        },
    }
    path = DATA / "clients_metadata.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2, ensure_ascii=False)


def main():
    print("Generating synthetic data for 3 clients...\n")
    all_stats: dict[str, list[tuple[str, int]]] = {}

    for name, gen_fn in [
        ("acme-logistics", gen_acme),
        ("transfroid-express", gen_transfroid),
        ("petit-colis", gen_petitcolis),
    ]:
        print(f"  [{name}]")
        stats = gen_fn()
        all_stats[name] = stats
        for filename, n in stats:
            print(f"    {filename:<55} {n:>6,} rows")
        print()

    write_metadata()

    grand_total = sum(n for stats in all_stats.values() for _, n in stats)
    total_files = sum(len(s) for s in all_stats.values()) + 1  # +1 for metadata
    print(f"{'=' * 65}")
    print(f"  {total_files} files, {grand_total:,} total rows in {DATA}")
    print(f"{'=' * 65}")


if __name__ == "__main__":
    main()
