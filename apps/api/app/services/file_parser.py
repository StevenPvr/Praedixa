"""File Parser — CSV and XLSX ingestion with French locale support.

Parses uploaded files into a list of row dicts with normalized values.
Detects format (csv, xlsx, lucca, payfit) via column header heuristics.

Security notes:
- Magic bytes validation rejects disguised binary files before parsing.
- Encoding detection uses a strict allowlist (utf-8, latin-1, cp1252,
  utf-8-sig). Any other detected encoding is rejected to prevent
  encoding confusion attacks (UTF-7, shift-JIS, etc.).
- max_rows is enforced during parsing to prevent memory exhaustion.
- Formula injection (=, +, -, @) is stripped from cell values to prevent
  CSV injection attacks when data is later exported.
- openpyxl is opened in read_only mode to avoid XML bomb / zip bomb DoS.
"""

from __future__ import annotations

import csv
import datetime
import io
import logging
import re
from dataclasses import dataclass, field
from typing import Any, Final

import chardet
from openpyxl import load_workbook

from app.core.config import settings

logger = logging.getLogger(__name__)

__all__ = ["ParseResult", "parse_file", "FileParseError"]

# ── Constants ────────────────────────────────────────────

# Strict encoding allowlist — any detection outside this set is rejected.
_ALLOWED_ENCODINGS: Final[frozenset[str]] = frozenset(
    {"utf-8", "latin-1", "iso-8859-1", "cp1252", "windows-1252", "utf-8-sig", "ascii"}
)

# Normalize chardet output to canonical Python codec names.
_ENCODING_NORMALIZATION: Final[dict[str, str]] = {
    "iso-8859-1": "latin-1",
    "windows-1252": "cp1252",
    "ascii": "utf-8",
}

# XLSX magic bytes: PK ZIP signature (50 4B 03 04).
_XLSX_MAGIC: Final[bytes] = b"PK"

# Delimiter detection priority order (French CSV convention: semicolon first).
_DELIMITER_CANDIDATES: Final[tuple[str, ...]] = (";", ",", "\t")

# ── Lucca detection: must contain >= 3 of these headers ──
_LUCCA_INDICATORS: Final[frozenset[str]] = frozenset(
    {
        "collaborateur",
        "matricule",
        "type de conge",
        "type d'absence",
        "date de debut",
        "date de fin",
        "duree",
        "service",
        "departement",
        "manager",
        "statut",
    }
)

# ── PayFit detection: must contain >= 3 of these headers ──
_PAYFIT_INDICATORS: Final[frozenset[str]] = frozenset(
    {
        "salaire brut",
        "salaire net",
        "net a payer",
        "hs 25%",
        "hs 50%",
        "heures supplementaires",
        "conges payes",
        "primes",
        "cotisations",
        "nom",
        "prenom",
        "matricule",
    }
)

_LUCCA_DETECTION_THRESHOLD: Final[int] = 3
_PAYFIT_DETECTION_THRESHOLD: Final[int] = 3

# French boolean mapping (case-insensitive).
_FRENCH_BOOLEANS: Final[dict[str, bool]] = {
    "oui": True,
    "non": False,
    "vrai": True,
    "faux": False,
    "true": True,
    "false": False,
    "1": True,
    "0": False,
    "o": True,
    "n": False,
}

# French date pattern: DD/MM/YYYY or DD-MM-YYYY
_FR_DATE_RE: Final[re.Pattern[str]] = re.compile(
    r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$"
)

# French decimal: "1 234,56" or "1234,56"
_FR_DECIMAL_RE: Final[re.Pattern[str]] = re.compile(r"^-?\d[\d \u00a0]*,\d+$")

# CSV formula injection prefixes — strip to prevent CSV injection.
_FORMULA_PREFIXES: Final[tuple[str, ...]] = ("=", "+", "-", "@")


# ── Exceptions ───────────────────────────────────────────


class FileParseError(Exception):
    """Raised when file parsing fails for a recoverable reason."""

    def __init__(self, message: str, *, code: str = "FILE_PARSE_ERROR") -> None:
        self.message = message
        self.code = code
        super().__init__(message)


# ── Result dataclass ─────────────────────────────────────


@dataclass(frozen=True)
class ParseResult:
    """Immutable result of file parsing."""

    rows: list[dict[str, Any]]
    source_columns: list[str]
    detected_format: str  # "csv" | "xlsx" | "lucca" | "payfit"
    detected_encoding: str
    warnings: list[str] = field(default_factory=list)
    row_count: int = 0


# ── Public API ───────────────────────────────────────────


def parse_file(
    content: bytes,
    filename: str,
    *,
    format_hint: str | None = None,
    sheet_name: str | None = None,
    max_rows: int | None = None,
) -> ParseResult:
    """Parse a CSV or XLSX file into a list of row dicts.

    Args:
        content: Raw file bytes.
        filename: Original filename (used for extension detection).
        format_hint: Optional format override ("csv", "xlsx", "lucca", "payfit").
        sheet_name: For XLSX files, which sheet to read (default: first sheet).
        max_rows: Maximum rows to parse. Defaults to settings.MAX_ROWS_PER_INGESTION.

    Returns:
        ParseResult with rows, source columns, detected format, and warnings.

    Raises:
        FileParseError: On invalid format, encoding, or structure.
    """
    if max_rows is None:
        max_rows = settings.MAX_ROWS_PER_INGESTION

    if not content:
        raise FileParseError("File is empty", code="EMPTY_FILE")

    ext = _extract_extension(filename)
    warnings: list[str] = []

    if ext == "xlsx" or format_hint == "xlsx":
        _validate_xlsx_magic(content)
        rows, columns = _parse_xlsx(content, sheet_name=sheet_name, max_rows=max_rows)
        encoding = "utf-8"
    elif ext == "csv" or format_hint in ("csv", "lucca", "payfit") or ext == "":
        encoding = _detect_encoding(content)
        text = _decode_content(content, encoding)
        rows, columns = _parse_csv(text, max_rows=max_rows)
    else:
        raise FileParseError(
            f"Unsupported file extension: .{ext}",
            code="UNSUPPORTED_FORMAT",
        )

    if not columns:
        raise FileParseError(
            "No columns detected in file",
            code="NO_COLUMNS",
        )

    if not rows:
        warnings.append("File contains headers but no data rows")

    # Detect format from column headers
    detected_format = _detect_format(columns, format_hint=format_hint, ext=ext)

    # Apply French coercion to all values
    coerced_rows = [_coerce_row(row, warnings) for row in rows]

    if len(coerced_rows) == max_rows:
        warnings.append(
            f"Row limit reached ({max_rows}). File may contain additional rows."
        )

    return ParseResult(
        rows=coerced_rows,
        source_columns=columns,
        detected_format=detected_format,
        detected_encoding=encoding,
        warnings=warnings,
        row_count=len(coerced_rows),
    )


# ── Private helpers ──────────────────────────────────────


def _extract_extension(filename: str) -> str:
    """Extract and normalize file extension (lowercase, no dot)."""
    name = filename.rsplit("/", 1)[-1]  # Handle path separators
    name = name.rsplit("\\", 1)[-1]  # Handle Windows paths
    if "." in name:
        return name.rsplit(".", 1)[-1].lower()
    return ""


def _validate_xlsx_magic(content: bytes) -> None:
    """Validate XLSX file starts with PK (ZIP) magic bytes."""
    _min_xlsx_size = 4
    if len(content) < _min_xlsx_size or content[:2] != _XLSX_MAGIC:
        raise FileParseError(
            "File does not have valid XLSX signature (expected PK/ZIP header)",
            code="INVALID_XLSX_SIGNATURE",
        )


def _detect_encoding(content: bytes) -> str:
    """Detect encoding via chardet, validated against allowlist.

    Returns a canonical Python codec name.
    Raises FileParseError for unsupported encodings.
    """
    # Check for UTF-8 BOM first
    if content[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"

    # Sample first 64KB for detection (sufficient and avoids slow detection)
    sample = content[:65536]
    result = chardet.detect(sample)
    detected = (result.get("encoding") or "").lower().strip()

    if not detected:
        raise FileParseError(
            "Could not detect file encoding",
            code="ENCODING_DETECTION_FAILED",
        )

    # Normalize to canonical names
    normalized = _ENCODING_NORMALIZATION.get(detected, detected)

    if normalized not in _ALLOWED_ENCODINGS:
        raise FileParseError(
            f"Unsupported encoding detected: {detected}. "
            "Allowed: UTF-8, Latin-1, CP1252, UTF-8 with BOM.",
            code="UNSUPPORTED_ENCODING",
        )

    return normalized


def _decode_content(content: bytes, encoding: str) -> str:
    """Decode bytes to string with the detected encoding."""
    try:
        return content.decode(encoding)
    except (UnicodeDecodeError, LookupError) as exc:
        raise FileParseError(
            f"Failed to decode file as {encoding}",
            code="DECODE_ERROR",
        ) from exc


def _detect_delimiter(text: str) -> str:
    """Detect CSV delimiter from first few lines.

    Uses frequency analysis on the first 10 lines. Returns the delimiter
    with the most consistent column count.
    """
    lines = text.split("\n", 20)[:10]
    lines = [ln for ln in lines if ln.strip()]

    if not lines:
        return ";"  # Default to French semicolon

    best_delimiter = ";"
    best_consistency = -1

    for delim in _DELIMITER_CANDIDATES:
        counts = [ln.count(delim) for ln in lines]
        if not counts or counts[0] == 0:
            continue
        # Consistency = how many lines have the same count as the header
        header_count = counts[0]
        consistency = sum(1 for c in counts if c == header_count)
        if consistency > best_consistency or (
            consistency == best_consistency and header_count > 0
        ):
            best_consistency = consistency
            best_delimiter = delim

    return best_delimiter


def _check_duplicate_columns(columns: list[str]) -> None:
    """Reject duplicate column headers (case-insensitive)."""
    seen: set[str] = set()
    for col in columns:
        lower = col.lower()
        if lower in seen:
            raise FileParseError(
                f"Duplicate column header detected: '{col}'",
                code="DUPLICATE_COLUMN",
            )
        seen.add(lower)


def _parse_csv(
    text: str,
    *,
    max_rows: int,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse CSV text into rows and column names.

    Uses stdlib csv module (not pandas) for minimal memory footprint.
    Enforces max_rows during iteration to prevent memory exhaustion.
    """
    delimiter = _detect_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter, strict=True)

    try:
        raw_headers = next(reader)
    except StopIteration as exc:
        raise FileParseError("CSV file has no header row", code="NO_HEADER") from exc

    # Strip BOM and whitespace from headers
    columns = [h.strip().strip("\ufeff") for h in raw_headers]

    _check_duplicate_columns(columns)

    # Parse rows up to max_rows
    rows: list[dict[str, Any]] = []
    for row_num, row in enumerate(reader, start=2):
        if row_num - 1 > max_rows:
            break

        # Skip completely empty rows
        if not any(cell.strip() for cell in row):
            continue

        # Build dict, padding or truncating to match header count
        row_dict: dict[str, Any] = {}
        for i, col in enumerate(columns):
            if i < len(row):
                value = row[i].strip()
                value = _sanitize_cell_value(value)
                row_dict[col] = value if value else None
            else:
                row_dict[col] = None
        rows.append(row_dict)

    return rows, columns


def _open_xlsx_worksheet(
    content: bytes,
    sheet_name: str | None,
) -> tuple[Any, Any]:
    """Open XLSX workbook and return (workbook, worksheet).

    Uses read_only mode to prevent XML bomb DoS.
    Raises FileParseError on invalid files or missing sheets.
    """
    try:
        wb = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise FileParseError(
            "Failed to open XLSX file. File may be corrupted or password-protected.",
            code="XLSX_OPEN_ERROR",
        ) from exc

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            # Do NOT reflect raw sheet_name in error — it comes from user input.
            # Only show available sheet names (these are from the file, not user).
            available = ", ".join(wb.sheetnames[:10])
            raise FileParseError(
                f"Requested sheet not found. Available: {available}",
                code="SHEET_NOT_FOUND",
            )
        return wb, wb[sheet_name]

    ws = wb.active
    if ws is None:
        wb.close()
        raise FileParseError(
            "Workbook has no active sheet",
            code="NO_ACTIVE_SHEET",
        )
    return wb, ws


def _read_xlsx_headers(ws: Any) -> list[str]:
    """Read and validate XLSX header row. Returns column names."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise FileParseError("XLSX sheet has no header row", code="NO_HEADER")

    columns = [
        str(h).strip() if h is not None else f"_unnamed_{i}"
        for i, h in enumerate(header_row)
    ]

    _check_duplicate_columns(columns)
    return columns


def _parse_xlsx(
    content: bytes,
    *,
    sheet_name: str | None = None,
    max_rows: int,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse XLSX content into rows and column names.

    Uses openpyxl in read_only mode to prevent XML bomb DoS
    and minimize memory consumption.
    """
    wb, ws = _open_xlsx_worksheet(content, sheet_name)
    try:
        columns = _read_xlsx_headers(ws)
        rows: list[dict[str, Any]] = []

        for row_values in ws.iter_rows(min_row=2, values_only=True):
            if len(rows) >= max_rows:
                break
            if not any(v is not None for v in row_values):
                continue

            row_dict: dict[str, Any] = {}
            for i, col in enumerate(columns):
                value = row_values[i] if i < len(row_values) else None
                if isinstance(value, str):
                    value = _sanitize_cell_value(value.strip())
                    value = value if value else None
                row_dict[col] = value
            rows.append(row_dict)
    finally:
        wb.close()

    return rows, columns


def _sanitize_cell_value(value: str) -> str:
    """Strip CSV formula injection prefixes from cell values.

    Prevents formula injection when data is later exported to spreadsheet
    formats. Strips leading =, +, @ characters. Leading '-' is only
    stripped if the remaining value does NOT look like a number (to
    preserve legitimate negative numeric values like "-42.5").
    """
    stripped = value
    while stripped and stripped[0] in ("=", "+", "@"):
        stripped = stripped[1:]
    # Strip leading '-' only if it's not a negative number.
    # A negative number starts with '-' followed by a digit.
    while stripped and stripped[0] == "-":
        rest = stripped[1:]
        if rest and rest[0].isdigit():
            # Looks like a negative number — preserve it
            break
        stripped = rest
    return stripped


def _detect_format(
    columns: list[str],
    *,
    format_hint: str | None,
    ext: str,
) -> str:
    """Detect file format from column headers.

    Priority:
    1. format_hint if provided and valid
    2. Lucca detection (>= 3 indicator columns)
    3. PayFit detection (>= 3 indicator columns)
    4. Extension-based ("xlsx" or "csv")
    """
    if format_hint in ("lucca", "payfit", "csv", "xlsx"):
        return format_hint

    normalized = {c.lower().strip() for c in columns}

    lucca_matches = len(normalized & _LUCCA_INDICATORS)
    if lucca_matches >= _LUCCA_DETECTION_THRESHOLD:
        return "lucca"

    payfit_matches = len(normalized & _PAYFIT_INDICATORS)
    if payfit_matches >= _PAYFIT_DETECTION_THRESHOLD:
        return "payfit"

    return "xlsx" if ext == "xlsx" else "csv"


def _coerce_row(
    row: dict[str, Any],
    warnings: list[str],
) -> dict[str, Any]:
    """Apply French locale coercion to all values in a row.

    Transforms:
    - Dates DD/MM/YYYY -> datetime.date
    - Decimals "1 234,56" -> float 1234.56
    - Booleans Oui/Non -> True/False
    """
    coerced: dict[str, Any] = {}
    for key, value in row.items():
        if value is None:
            coerced[key] = None
            continue
        if not isinstance(value, str):
            # XLSX may return native types (int, float, datetime)
            coerced[key] = value
            continue
        coerced[key] = _coerce_value(value, key, warnings)
    return coerced


def _coerce_value(
    value: str,
    column_name: str,
    warnings: list[str],
) -> Any:
    """Attempt French-locale coercion on a single string value.

    Tries in order: date, decimal, boolean. Falls back to original string.
    """
    # 1. French date: DD/MM/YYYY
    date_match = _FR_DATE_RE.match(value)
    if date_match:
        day, month, year = (
            int(date_match.group(1)),
            int(date_match.group(2)),
            int(date_match.group(3)),
        )
        try:
            return datetime.date(year, month, day)
        except ValueError:
            warnings.append(
                f"Invalid date value in column '{column_name}': "
                f"parsed as {day}/{month}/{year} but date is out of range"
            )
            return value

    # 2. French decimal: "1 234,56" or "1234,56"
    if _FR_DECIMAL_RE.match(value):
        try:
            cleaned = value.replace(" ", "").replace("\u00a0", "").replace(",", ".")
            return float(cleaned)
        except ValueError:
            return value

    # 3. French boolean
    lower = value.lower().strip()
    if lower in _FRENCH_BOOLEANS:
        return _FRENCH_BOOLEANS[lower]

    return value
