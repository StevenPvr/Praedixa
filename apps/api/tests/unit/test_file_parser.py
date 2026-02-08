"""Tests for File Parser service.

Covers:
- CSV parsing: semicolon, comma, tab delimiters
- Encoding detection: utf-8, latin-1, cp1252, utf-8-bom, unsupported
- Encoding normalization: iso-8859-1 -> latin-1, windows-1252 -> cp1252, ascii -> utf-8
- French date coercion: DD/MM/YYYY -> datetime.date, single-digit, leap years, invalid
- French decimal coercion: "1 234,56" -> 1234.56, negative, NBSP, zero
- French boolean coercion: Oui/Non/Vrai/Faux/O/N/True/False/1/0
- XLSX parsing via openpyxl in read_only mode
- Lucca format detection from column headers
- PayFit format detection from column headers
- Magic bytes validation (reject non-XLSX binary)
- max_rows enforcement (from arg and from settings)
- Empty file / no header / no data handling
- Duplicate column detection (case-insensitive)
- Formula injection stripping: =, +, @, and '-' with negative-number preservation
- Sheet name selection / missing sheet / no active sheet
- _coerce_row: None values, non-string native types, string coercion
- _decode_content: UnicodeDecodeError and LookupError
- Extension extraction edge cases (paths, dots, empty)
- ParseResult / FileParseError dataclass fields
"""

import datetime
import io
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from app.services.file_parser import (
    FileParseError,
    ParseResult,
    _check_duplicate_columns,
    _coerce_row,
    _coerce_value,
    _decode_content,
    _detect_delimiter,
    _detect_encoding,
    _detect_format,
    _extract_extension,
    _open_xlsx_worksheet,
    _parse_csv,
    _parse_xlsx,
    _read_xlsx_headers,
    _sanitize_cell_value,
    _validate_xlsx_magic,
    parse_file,
)

# ── Fixtures ─────────────────────────────────────────────


def _make_csv_bytes(text: str, encoding: str = "utf-8") -> bytes:
    """Encode CSV text to bytes with given encoding."""
    return text.encode(encoding)


def _make_xlsx_bytes(
    headers: list[str],
    rows: list[list],
    sheet_name: str = "Sheet1",
) -> bytes:
    """Create an XLSX file in-memory and return bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── _extract_extension ───────────────────────────────────


class TestExtractExtension:
    def test_csv_extension(self) -> None:
        assert _extract_extension("data.csv") == "csv"

    def test_xlsx_extension(self) -> None:
        assert _extract_extension("report.xlsx") == "xlsx"

    def test_uppercase_extension(self) -> None:
        assert _extract_extension("DATA.CSV") == "csv"

    def test_no_extension(self) -> None:
        assert _extract_extension("myfile") == ""

    def test_path_with_dirs(self) -> None:
        assert _extract_extension("/var/uploads/data.csv") == "csv"

    def test_windows_path(self) -> None:
        assert _extract_extension("C:\\Users\\data.xlsx") == "xlsx"

    def test_multiple_dots(self) -> None:
        assert _extract_extension("my.data.file.csv") == "csv"

    def test_empty_filename(self) -> None:
        assert _extract_extension("") == ""

    def test_dot_only(self) -> None:
        assert _extract_extension(".") == ""

    def test_hidden_file(self) -> None:
        assert _extract_extension(".gitignore") == "gitignore"


# ── _validate_xlsx_magic ─────────────────────────────────


class TestValidateXlsxMagic:
    def test_valid_xlsx_magic(self) -> None:
        xlsx = _make_xlsx_bytes(["col"], [[1]])
        _validate_xlsx_magic(xlsx)  # Should not raise

    def test_csv_content_rejected(self) -> None:
        with pytest.raises(FileParseError, match="XLSX signature") as exc_info:
            _validate_xlsx_magic(b"col1;col2\n1;2\n")
        assert exc_info.value.code == "INVALID_XLSX_SIGNATURE"

    def test_empty_content_rejected(self) -> None:
        with pytest.raises(FileParseError, match="XLSX signature"):
            _validate_xlsx_magic(b"")

    def test_short_content_rejected(self) -> None:
        with pytest.raises(FileParseError, match="XLSX signature"):
            _validate_xlsx_magic(b"PK")  # Only 2 bytes, need 4

    def test_three_bytes_rejected(self) -> None:
        with pytest.raises(FileParseError, match="XLSX signature"):
            _validate_xlsx_magic(b"PK\x03")  # 3 bytes < 4 min

    def test_random_binary_rejected(self) -> None:
        with pytest.raises(FileParseError, match="XLSX signature"):
            _validate_xlsx_magic(b"\x89PNG\r\n\x1a\n")  # PNG header

    def test_four_bytes_with_pk_passes(self) -> None:
        _validate_xlsx_magic(b"PK\x03\x04")  # 4 bytes, starts with PK


# ── _detect_encoding ─────────────────────────────────────


class TestDetectEncoding:
    def test_utf8(self) -> None:
        content = b"nom;prenom\nDupont;Jean\n"
        enc = _detect_encoding(content)
        assert enc in ("utf-8", "ascii")

    def test_utf8_bom(self) -> None:
        content = b"\xef\xbb\xbfnom;prenom\n"
        assert _detect_encoding(content) == "utf-8-sig"

    def test_latin1(self) -> None:
        content = "nom;pr\xe9nom;r\xe9sum\xe9\n".encode("latin-1")
        enc = _detect_encoding(content)
        assert enc in ("latin-1", "cp1252", "utf-8")

    def test_empty_content_raises(self) -> None:
        with pytest.raises(FileParseError, match="encoding"):
            _detect_encoding(b"")

    def test_unsupported_encoding_detected(self) -> None:
        fake_result = {"encoding": "shift_jis", "confidence": 0.99}
        with patch("app.services.file_parser.chardet.detect", return_value=fake_result):
            with pytest.raises(
                FileParseError, match="Unsupported encoding"
            ) as exc_info:
                _detect_encoding(b"some data bytes")
            assert exc_info.value.code == "UNSUPPORTED_ENCODING"

    def test_none_encoding_raises(self) -> None:
        fake_result = {"encoding": None, "confidence": 0}
        with patch("app.services.file_parser.chardet.detect", return_value=fake_result):
            with pytest.raises(FileParseError, match="Could not detect") as exc_info:
                _detect_encoding(b"bytes")
            assert exc_info.value.code == "ENCODING_DETECTION_FAILED"

    def test_empty_string_encoding_raises(self) -> None:
        fake_result = {"encoding": "", "confidence": 0}
        with (
            patch("app.services.file_parser.chardet.detect", return_value=fake_result),
            pytest.raises(FileParseError, match="Could not detect"),
        ):
            _detect_encoding(b"bytes")

    def test_iso_8859_1_normalizes_to_latin1(self) -> None:
        fake_result = {"encoding": "ISO-8859-1", "confidence": 0.99}
        with patch("app.services.file_parser.chardet.detect", return_value=fake_result):
            assert _detect_encoding(b"some bytes") == "latin-1"

    def test_windows_1252_normalizes_to_cp1252(self) -> None:
        fake_result = {"encoding": "Windows-1252", "confidence": 0.99}
        with patch("app.services.file_parser.chardet.detect", return_value=fake_result):
            assert _detect_encoding(b"some bytes") == "cp1252"

    def test_ascii_normalizes_to_utf8(self) -> None:
        fake_result = {"encoding": "ascii", "confidence": 0.99}
        with patch("app.services.file_parser.chardet.detect", return_value=fake_result):
            assert _detect_encoding(b"some bytes") == "utf-8"

    def test_utf7_rejected(self) -> None:
        fake_result = {"encoding": "UTF-7", "confidence": 0.90}
        with (
            patch("app.services.file_parser.chardet.detect", return_value=fake_result),
            pytest.raises(FileParseError, match="Unsupported encoding"),
        ):
            _detect_encoding(b"some bytes")


# ── _decode_content ──────────────────────────────────────


class TestDecodeContent:
    def test_utf8_decode(self) -> None:
        content = b"Bonjour"
        assert _decode_content(content, "utf-8") == "Bonjour"

    def test_latin1_decode(self) -> None:
        content = "résumé".encode("latin-1")
        assert _decode_content(content, "latin-1") == "résumé"

    def test_cp1252_decode(self) -> None:
        content = b"\x93hello\x94"  # smart quotes in CP1252
        result = _decode_content(content, "cp1252")
        assert "hello" in result

    def test_unicode_error_raises(self) -> None:
        with pytest.raises(FileParseError, match="Failed to decode") as exc_info:
            _decode_content(b"\x80\x81\x82", "utf-8")
        assert exc_info.value.code == "DECODE_ERROR"

    def test_lookup_error_raises(self) -> None:
        with pytest.raises(FileParseError, match="Failed to decode") as exc_info:
            _decode_content(b"hello", "not-a-real-codec")
        assert exc_info.value.code == "DECODE_ERROR"


# ── _detect_delimiter ────────────────────────────────────


class TestDetectDelimiter:
    def test_semicolon(self) -> None:
        text = "a;b;c\n1;2;3\n4;5;6\n"
        assert _detect_delimiter(text) == ";"

    def test_comma(self) -> None:
        text = "a,b,c\n1,2,3\n4,5,6\n"
        assert _detect_delimiter(text) == ","

    def test_tab(self) -> None:
        text = "a\tb\tc\n1\t2\t3\n"
        assert _detect_delimiter(text) == "\t"

    def test_equal_consistency_last_delimiter_wins(self) -> None:
        """When all delimiters have equal consistency, the last in
        _DELIMITER_CANDIDATES with header_count > 0 wins."""
        # Both ; and , have consistency=3, comma tested after semicolon -> comma wins
        text = "a;b,c\n1;2,3\n4;5,6\n"
        assert _detect_delimiter(text) == ","

    def test_semicolon_wins_with_higher_header_count(self) -> None:
        """Semicolon wins when it appears more times per line (higher header_count)
        — but only if its consistency is strictly greater than comma's."""
        # 3 semicolons per line, 0 commas -> semicolon wins unambiguously
        text = "a;b;c;d\n1;2;3;4\n5;6;7;8\n"
        assert _detect_delimiter(text) == ";"

    def test_empty_text(self) -> None:
        assert _detect_delimiter("") == ";"

    def test_whitespace_only_lines(self) -> None:
        assert _detect_delimiter("   \n   \n") == ";"

    def test_single_line(self) -> None:
        text = "a;b;c\n"
        assert _detect_delimiter(text) == ";"

    def test_no_delimiters_at_all(self) -> None:
        text = "name\nAlice\nBob"
        assert _detect_delimiter(text) == ";"

    def test_more_than_10_lines_uses_first_10(self) -> None:
        lines = ["a;b;c"] + [f"{i};{i};{i}" for i in range(25)]
        text = "\n".join(lines)
        assert _detect_delimiter(text) == ";"


# ── _sanitize_cell_value ─────────────────────────────────


class TestSanitizeCellValue:
    def test_normal_value(self) -> None:
        assert _sanitize_cell_value("hello") == "hello"

    def test_formula_equals(self) -> None:
        assert _sanitize_cell_value("=CMD('calc')") == "CMD('calc')"

    def test_formula_plus(self) -> None:
        assert _sanitize_cell_value("+1234") == "1234"

    def test_formula_at(self) -> None:
        assert _sanitize_cell_value("@SUM(A1)") == "SUM(A1)"

    def test_multiple_prefixes(self) -> None:
        """The function strips =,+,@ in one pass, then '-' in a second pass.
        '=-+@evil' -> strip '=' -> '-+@evil' -> '-' rest is '+@evil',
        '+' is not digit so strip '-' -> '+@evil'. The + is NOT re-stripped
        because the first loop (=,+,@) already finished."""
        assert _sanitize_cell_value("=-+@evil") == "+@evil"

    def test_empty_string(self) -> None:
        assert _sanitize_cell_value("") == ""

    def test_negative_number_preserved(self) -> None:
        """Leading '-' followed by digit is preserved (negative number)."""
        assert _sanitize_cell_value("-42") == "-42"
        assert _sanitize_cell_value("-3.14") == "-3.14"
        assert _sanitize_cell_value("-0") == "-0"

    def test_minus_non_number_stripped(self) -> None:
        """Leading '-' NOT followed by digit is stripped (formula injection)."""
        assert _sanitize_cell_value("-cmd") == "cmd"

    def test_minus_empty_rest_stripped(self) -> None:
        """Single '-' with nothing after -> stripped to empty."""
        assert _sanitize_cell_value("-") == ""

    def test_double_minus_non_number(self) -> None:
        """--evil -> evil (both dashes stripped since first rest is '-evil',
        '-' is not a digit, so strip; then 'evil' starts with 'e', strip again? No:
        after first strip, rest='evil', 'e'[0] != '-', so loop ends.)"""
        assert _sanitize_cell_value("--evil") == "evil"

    def test_equals_then_minus_number(self) -> None:
        """=-42 -> first loop strips '=', then '-42' -> '-' followed by '4' (digit),
        so preserved."""
        assert _sanitize_cell_value("=-42") == "-42"

    def test_prefix_in_middle_not_stripped(self) -> None:
        assert _sanitize_cell_value("a=b+c") == "a=b+c"

    def test_only_formula_chars_no_minus(self) -> None:
        assert _sanitize_cell_value("=+@") == ""


# ── _check_duplicate_columns ─────────────────────────────


class TestCheckDuplicateColumns:
    def test_no_duplicates(self) -> None:
        _check_duplicate_columns(["a", "b", "c"])  # No raise

    def test_exact_duplicate(self) -> None:
        with pytest.raises(FileParseError, match="Duplicate") as exc_info:
            _check_duplicate_columns(["a", "b", "a"])
        assert exc_info.value.code == "DUPLICATE_COLUMN"

    def test_case_insensitive_duplicate(self) -> None:
        with pytest.raises(FileParseError, match="Duplicate"):
            _check_duplicate_columns(["Name", "name"])

    def test_empty_list(self) -> None:
        _check_duplicate_columns([])  # No raise

    def test_single_column(self) -> None:
        _check_duplicate_columns(["only"])  # No raise


# ── French coercion ──────────────────────────────────────


class TestCoerceValue:
    def test_french_date_slash(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("25/12/2025", "date_col", warnings)
        assert result == datetime.date(2025, 12, 25)
        assert len(warnings) == 0

    def test_french_date_dash(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("01-06-2024", "date_col", warnings)
        assert result == datetime.date(2024, 6, 1)

    def test_french_date_single_digit(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("1/1/2000", "date_col", warnings)
        assert result == datetime.date(2000, 1, 1)

    def test_french_date_invalid_day(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("31/02/2025", "date_col", warnings)
        assert result == "31/02/2025"
        assert len(warnings) == 1
        assert "out of range" in warnings[0]

    def test_french_date_invalid_month(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("15/13/2025", "col", warnings)
        assert result == "15/13/2025"
        assert len(warnings) == 1

    def test_french_date_feb_29_leap_year(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("29/02/2024", "col", warnings)
        assert result == datetime.date(2024, 2, 29)

    def test_french_date_feb_29_non_leap(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("29/02/2023", "col", warnings)
        assert result == "29/02/2023"
        assert len(warnings) == 1

    def test_french_decimal(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("1 234,56", "amount", warnings)
        assert result == pytest.approx(1234.56)

    def test_french_decimal_simple(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("42,5", "rate", warnings)
        assert result == pytest.approx(42.5)

    def test_french_decimal_negative(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("-100,00", "diff", warnings)
        assert result == pytest.approx(-100.0)

    def test_french_decimal_zero(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("0,00", "val", warnings)
        assert result == pytest.approx(0.0)

    def test_french_decimal_with_nbsp(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("1\u00a0234,56", "val", warnings)
        assert result == pytest.approx(1234.56)

    def test_french_decimal_large(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("1 000 000,99", "val", warnings)
        assert result == pytest.approx(1000000.99)

    def test_french_decimal_value_error_fallback(self) -> None:
        """Regex matches but float() fails — returns original string."""
        warnings: list[str] = []
        with patch("app.services.file_parser.float", side_effect=ValueError("forced")):
            result = _coerce_value("1 234,56", "amount", warnings)
        assert result == "1 234,56"

    @pytest.mark.parametrize(
        ("input_val", "expected"),
        [
            ("oui", True),
            ("non", False),
            ("vrai", True),
            ("faux", False),
            ("true", True),
            ("false", False),
            ("1", True),
            ("0", False),
            ("o", True),
            ("n", False),
            ("OUI", True),
            ("NON", False),
            ("Vrai", True),
            ("Faux", False),
        ],
    )
    def test_french_boolean(self, input_val, expected) -> None:
        warnings: list[str] = []
        result = _coerce_value(input_val, "col", warnings)
        assert result is expected

    def test_plain_string(self) -> None:
        warnings: list[str] = []
        assert _coerce_value("Jean Dupont", "name", warnings) == "Jean Dupont"

    def test_numeric_string_not_french_decimal(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("1234.56", "amount", warnings)
        assert result == "1234.56"

    def test_iso_date_not_matched(self) -> None:
        warnings: list[str] = []
        result = _coerce_value("2025-01-15", "col", warnings)
        assert result == "2025-01-15"


# ── _coerce_row ──────────────────────────────────────────


class TestCoerceRow:
    def test_none_values_preserved(self) -> None:
        warnings: list[str] = []
        result = _coerce_row({"a": None, "b": "hello"}, warnings)
        assert result["a"] is None
        assert result["b"] == "hello"

    def test_native_int_preserved(self) -> None:
        warnings: list[str] = []
        result = _coerce_row({"age": 30}, warnings)
        assert result["age"] == 30

    def test_native_float_preserved(self) -> None:
        warnings: list[str] = []
        result = _coerce_row({"score": 95.5}, warnings)
        assert result["score"] == 95.5

    def test_native_datetime_preserved(self) -> None:
        dt = datetime.datetime(2025, 1, 15, 10, 30, 0)  # noqa: DTZ001 -- naive datetime to test coerce_row preserves native datetimes
        warnings: list[str] = []
        result = _coerce_row({"ts": dt}, warnings)
        assert result["ts"] is dt

    def test_string_values_coerced(self) -> None:
        warnings: list[str] = []
        result = _coerce_row({"date": "15/03/2025", "active": "oui"}, warnings)
        assert result["date"] == datetime.date(2025, 3, 15)
        assert result["active"] is True

    def test_mixed_types_in_one_row(self) -> None:
        warnings: list[str] = []
        result = _coerce_row(
            {"s": "text", "n": None, "i": 42, "d": "01/01/2000"},
            warnings,
        )
        assert result["s"] == "text"
        assert result["n"] is None
        assert result["i"] == 42
        assert result["d"] == datetime.date(2000, 1, 1)


# ── _detect_format ───────────────────────────────────────


class TestDetectFormat:
    def test_format_hint_csv(self) -> None:
        assert _detect_format(["a"], format_hint="csv", ext="csv") == "csv"

    def test_format_hint_xlsx(self) -> None:
        assert _detect_format(["a"], format_hint="xlsx", ext="xlsx") == "xlsx"

    def test_format_hint_lucca(self) -> None:
        assert _detect_format(["a"], format_hint="lucca", ext="csv") == "lucca"

    def test_format_hint_payfit(self) -> None:
        assert _detect_format(["a"], format_hint="payfit", ext="csv") == "payfit"

    def test_format_hint_unknown_falls_through(self) -> None:
        result = _detect_format(["random_col"], format_hint="unknown", ext="csv")
        assert result == "csv"

    def test_format_hint_none_falls_through(self) -> None:
        result = _detect_format(["random_col"], format_hint=None, ext="csv")
        assert result == "csv"

    def test_lucca_detection(self) -> None:
        cols = ["Collaborateur", "Matricule", "Type de conge", "Duree"]
        assert _detect_format(cols, format_hint=None, ext="csv") == "lucca"

    def test_lucca_exactly_threshold(self) -> None:
        cols = ["collaborateur", "matricule", "type de conge"]
        assert _detect_format(cols, format_hint=None, ext="csv") == "lucca"

    def test_lucca_below_threshold(self) -> None:
        cols = ["collaborateur", "matricule", "other"]
        assert _detect_format(cols, format_hint=None, ext="csv") == "csv"

    def test_payfit_detection(self) -> None:
        cols = ["Salaire brut", "Net a payer", "Cotisations", "Nom"]
        assert _detect_format(cols, format_hint=None, ext="csv") == "payfit"

    def test_payfit_below_threshold(self) -> None:
        cols = ["Salaire brut", "other"]
        assert _detect_format(cols, format_hint=None, ext="csv") == "csv"

    def test_lucca_priority_over_payfit(self) -> None:
        cols = [
            "collaborateur",
            "matricule",
            "type de conge",
            "salaire brut",
            "salaire net",
            "cotisations",
        ]
        assert _detect_format(cols, format_hint=None, ext="csv") == "lucca"

    def test_xlsx_fallback(self) -> None:
        assert _detect_format(["random"], format_hint=None, ext="xlsx") == "xlsx"

    def test_csv_fallback_empty_ext(self) -> None:
        assert _detect_format(["random"], format_hint=None, ext="") == "csv"

    def test_case_insensitive_indicator(self) -> None:
        cols = ["COLLABORATEUR", "Matricule", "Type De Conge"]
        assert _detect_format(cols, format_hint=None, ext="csv") == "lucca"


# ── _parse_csv ───────────────────────────────────────────


class TestParseCsv:
    def test_basic_csv(self) -> None:
        text = "nom;prenom;age\nDupont;Jean;30\nMartin;Marie;25\n"
        rows, cols = _parse_csv(text, max_rows=1000)
        assert cols == ["nom", "prenom", "age"]
        assert len(rows) == 2
        assert rows[0]["nom"] == "Dupont"

    def test_max_rows_enforcement(self) -> None:
        lines = ["a;b"] + [f"v{i};v{i}" for i in range(100)]
        text = "\n".join(lines)
        rows, _ = _parse_csv(text, max_rows=5)
        assert len(rows) == 5

    def test_empty_rows_skipped(self) -> None:
        text = "nom;age\nAlice;30\n  ;  \nBob;25\n"
        rows, _ = _parse_csv(text, max_rows=1000)
        assert len(rows) == 2

    def test_no_header_raises(self) -> None:
        with pytest.raises(FileParseError, match="no header") as exc_info:
            _parse_csv("", max_rows=1000)
        assert exc_info.value.code == "NO_HEADER"

    def test_header_only_no_data(self) -> None:
        text = "nom;prenom;age\n"
        rows, cols = _parse_csv(text, max_rows=1000)
        assert cols == ["nom", "prenom", "age"]
        assert len(rows) == 0

    def test_formula_injection_stripped(self) -> None:
        text = "name;formula\nAlice;=SUM(A1)\n"
        rows, _ = _parse_csv(text, max_rows=1000)
        assert rows[0]["formula"] == "SUM(A1)"

    def test_bom_stripped_from_headers(self) -> None:
        text = "\ufeffnom;prenom\nDupont;Jean\n"
        _rows, cols = _parse_csv(text, max_rows=1000)
        assert cols == ["nom", "prenom"]

    def test_short_row_padded_with_none(self) -> None:
        text = "a;b;c\n1;2\n"
        rows, _ = _parse_csv(text, max_rows=1000)
        assert rows[0]["c"] is None

    def test_empty_cell_becomes_none(self) -> None:
        text = "a;b\nvalue;\n"
        rows, _ = _parse_csv(text, max_rows=1000)
        assert rows[0]["b"] is None

    def test_whitespace_cell_becomes_none(self) -> None:
        text = "a;b\nvalue;   \n"
        rows, _ = _parse_csv(text, max_rows=1000)
        assert rows[0]["b"] is None

    def test_comma_delimiter(self) -> None:
        text = "name,age\nAlice,30\n"
        rows, cols = _parse_csv(text, max_rows=1000)
        assert cols == ["name", "age"]
        assert rows[0]["name"] == "Alice"

    def test_tab_delimiter(self) -> None:
        text = "name\tage\nAlice\t30\n"
        _rows, cols = _parse_csv(text, max_rows=1000)
        assert cols == ["name", "age"]

    def test_duplicate_headers_raises(self) -> None:
        text = "nom;nom;age\nA;B;30\n"
        with pytest.raises(FileParseError, match="Duplicate"):
            _parse_csv(text, max_rows=1000)


# ── _parse_xlsx ──────────────────────────────────────────


class TestParseXlsx:
    def test_basic_xlsx(self) -> None:
        content = _make_xlsx_bytes(
            ["nom", "age"],
            [["Dupont", 42], ["Martin", 35]],
        )
        rows, cols = _parse_xlsx(content, max_rows=1000)
        assert cols == ["nom", "age"]
        assert len(rows) == 2
        assert rows[0]["nom"] == "Dupont"
        assert rows[0]["age"] == 42

    def test_max_rows_enforcement(self) -> None:
        rows_data = [[i] for i in range(100)]
        content = _make_xlsx_bytes(["val"], rows_data)
        rows, _ = _parse_xlsx(content, max_rows=10)
        assert len(rows) == 10

    def test_empty_rows_skipped(self) -> None:
        content = _make_xlsx_bytes(["val"], [[1], [None], [3]])
        rows, _ = _parse_xlsx(content, max_rows=1000)
        assert len(rows) == 2

    def test_sheet_name_selection(self) -> None:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Metadata"
        ws1.append(["info"])
        ws1.append(["v1"])
        ws2 = wb.create_sheet("Data")
        ws2.append(["nom", "age"])
        ws2.append(["Dupont", 42])
        buf = io.BytesIO()
        wb.save(buf)
        rows, cols = _parse_xlsx(buf.getvalue(), sheet_name="Data", max_rows=1000)
        assert cols == ["nom", "age"]
        assert rows[0]["nom"] == "Dupont"

    def test_missing_sheet_raises(self) -> None:
        content = _make_xlsx_bytes(["col"], [[1]])
        with pytest.raises(FileParseError, match="not found") as exc_info:
            _parse_xlsx(content, sheet_name="NonExistent", max_rows=1000)
        assert exc_info.value.code == "SHEET_NOT_FOUND"

    def test_corrupt_xlsx_raises(self) -> None:
        with pytest.raises(FileParseError, match="Failed to open") as exc_info:
            _parse_xlsx(b"PK\x03\x04" + b"\x00" * 100, max_rows=1000)
        assert exc_info.value.code == "XLSX_OPEN_ERROR"

    def test_none_header_becomes_unnamed(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["name", None, "age"])
        ws.append(["Alice", "extra", 30])
        buf = io.BytesIO()
        wb.save(buf)
        _rows, cols = _parse_xlsx(buf.getvalue(), max_rows=1000)
        assert cols[1] == "_unnamed_1"

    def test_string_formula_sanitized(self) -> None:
        content = _make_xlsx_bytes(["val"], [["=evil"]])
        result_rows, _ = _parse_xlsx(content, max_rows=1000)
        # openpyxl data_only=True returns None for formula cells without cache
        if result_rows:
            assert result_rows[0]["val"] in (None, "evil")

    def test_empty_string_cell_becomes_none(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["col_a"])
        ws.append(["   "])
        buf = io.BytesIO()
        wb.save(buf)
        rows, _ = _parse_xlsx(buf.getvalue(), max_rows=1000)
        if rows:
            assert rows[0]["col_a"] is None

    def test_short_row_padded(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["a", "b", "c"])
        ws.append(["val"])
        buf = io.BytesIO()
        wb.save(buf)
        rows, _ = _parse_xlsx(buf.getvalue(), max_rows=1000)
        assert rows[0]["b"] is None
        assert rows[0]["c"] is None


# ── _read_xlsx_headers ───────────────────────────────────


class TestReadXlsxHeaders:
    def test_no_header_row_raises(self) -> None:
        class FakeWS:
            def iter_rows(self, **_kw):
                return iter([])

        with pytest.raises(FileParseError, match="no header") as exc_info:
            _read_xlsx_headers(FakeWS())
        assert exc_info.value.code == "NO_HEADER"

    def test_duplicate_headers_raises(self) -> None:
        ws = MagicMock()
        ws.iter_rows.return_value = iter([("nom", "prenom", "nom")])
        with pytest.raises(FileParseError, match="Duplicate"):
            _read_xlsx_headers(ws)


# ── _open_xlsx_worksheet ──────────────────────────────────


class TestOpenXlsxWorksheet:
    def test_no_active_sheet_raises(self) -> None:
        wb = Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        with patch("app.services.file_parser.load_workbook") as mock_lwb:
            mock_wb = mock_lwb.return_value
            mock_wb.active = None
            mock_wb.sheetnames = ["Sheet1"]
            with pytest.raises(FileParseError, match="no active sheet") as exc_info:
                _open_xlsx_worksheet(content, None)
            assert exc_info.value.code == "NO_ACTIVE_SHEET"
            mock_wb.close.assert_called_once()

    def test_missing_sheet_lists_available(self) -> None:
        content = _make_xlsx_bytes(["col"], [[1]])
        with pytest.raises(FileParseError, match=r"Available.*Sheet1"):
            _open_xlsx_worksheet(content, "NoSuch")


# ── parse_file — CSV ─────────────────────────────────────


class TestParseFileCsv:
    def test_basic_csv_semicolon(self) -> None:
        content = _make_csv_bytes("nom;age\nDupont;42\nMartin;35\n")
        result = parse_file(content, "data.csv")
        assert isinstance(result, ParseResult)
        assert result.row_count == 2
        assert result.detected_format == "csv"
        assert len(result.source_columns) == 2
        assert result.rows[0]["nom"] == "Dupont"

    def test_basic_csv_comma(self) -> None:
        content = _make_csv_bytes("name,age\nAlice,30\nBob,25\n")
        result = parse_file(content, "data.csv")
        assert result.row_count == 2
        assert result.rows[0]["name"] == "Alice"

    def test_csv_tab_delimiter(self) -> None:
        content = _make_csv_bytes("nom\tage\nDupont\t42\n")
        result = parse_file(content, "data.csv")
        assert result.row_count == 1

    def test_csv_french_coercion(self) -> None:
        content = _make_csv_bytes("date;montant;actif\n25/12/2025;1 234,56;Oui\n")
        result = parse_file(content, "data.csv")
        row = result.rows[0]
        assert row["date"] == datetime.date(2025, 12, 25)
        assert row["montant"] == pytest.approx(1234.56)
        assert row["actif"] is True

    def test_csv_empty_file(self) -> None:
        with pytest.raises(FileParseError, match="empty") as exc_info:
            parse_file(b"", "data.csv")
        assert exc_info.value.code == "EMPTY_FILE"

    def test_csv_only_newlines(self) -> None:
        """A file of only newlines: csv.reader reads first line as [],
        which produces columns=[], then parse_file raises NO_COLUMNS."""
        with pytest.raises(FileParseError, match="No columns"):
            parse_file(b"\n\n\n", "data.csv")

    def test_csv_header_only(self) -> None:
        content = _make_csv_bytes("nom;age\n")
        result = parse_file(content, "data.csv")
        assert result.row_count == 0
        assert "no data" in result.warnings[0].lower()

    def test_csv_max_rows_explicit(self) -> None:
        lines = ["val\n"] + [f"{i}\n" for i in range(100)]
        content = _make_csv_bytes("".join(lines))
        result = parse_file(content, "data.csv", max_rows=10)
        assert result.row_count == 10
        assert any("limit" in w.lower() for w in result.warnings)

    def test_csv_max_rows_from_settings(self) -> None:
        lines = ["a;b\n"] + [f"{i};{i}\n" for i in range(20)]
        content = _make_csv_bytes("".join(lines))
        with patch("app.services.file_parser.settings") as mock_settings:
            mock_settings.MAX_ROWS_PER_INGESTION = 5
            result = parse_file(content, "data.csv")
        assert result.row_count == 5

    def test_csv_duplicate_columns(self) -> None:
        content = _make_csv_bytes("nom;nom\nA;B\n")
        with pytest.raises(FileParseError, match="Duplicate"):
            parse_file(content, "data.csv")

    def test_csv_formula_injection_stripped(self) -> None:
        content = _make_csv_bytes("val\n=CMD('calc')\n+1234\n")
        result = parse_file(content, "data.csv")
        assert result.rows[0]["val"] == "CMD('calc')"
        assert result.rows[1]["val"] == "1234"

    def test_csv_encoding_latin1(self) -> None:
        content = "nom;prénom\nDupont;Hélène\n".encode("latin-1")
        result = parse_file(content, "data.csv")
        assert result.row_count == 1

    def test_csv_encoding_utf8_bom(self) -> None:
        content = b"\xef\xbb\xbfnom;age\nDupont;42\n"
        result = parse_file(content, "data.csv")
        assert result.row_count == 1
        assert "nom" in result.source_columns
        assert result.detected_encoding == "utf-8-sig"

    def test_csv_empty_rows_skipped(self) -> None:
        content = _make_csv_bytes("val\n1\n\n\n2\n")
        result = parse_file(content, "data.csv")
        assert result.row_count == 2

    def test_csv_lucca_format_detection(self) -> None:
        content = _make_csv_bytes(
            "Collaborateur;Matricule;Type de conge;Duree\nDupont;001;RTT;2\n"
        )
        result = parse_file(content, "data.csv")
        assert result.detected_format == "lucca"

    def test_csv_payfit_format_detection(self) -> None:
        content = _make_csv_bytes(
            "Nom;Salaire brut;Net a payer;Cotisations\nDupont;3500;2800;700\n"
        )
        result = parse_file(content, "data.csv")
        assert result.detected_format == "payfit"

    def test_csv_none_for_empty_cells(self) -> None:
        content = _make_csv_bytes("a;b;c\n1;;3\n")
        result = parse_file(content, "data.csv")
        assert result.rows[0]["b"] is None

    def test_csv_with_fewer_cols_than_header(self) -> None:
        content = _make_csv_bytes("a;b;c\n1;2\n")
        result = parse_file(content, "data.csv")
        assert result.rows[0]["c"] is None


# ── parse_file — XLSX ────────────────────────────────────


class TestParseFileXlsx:
    def test_basic_xlsx(self) -> None:
        content = _make_xlsx_bytes(
            ["nom", "age"],
            [["Dupont", 42], ["Martin", 35]],
        )
        result = parse_file(content, "data.xlsx")
        assert result.row_count == 2
        assert result.detected_format == "xlsx"
        assert result.detected_encoding == "utf-8"
        assert result.rows[0]["nom"] == "Dupont"
        assert result.rows[0]["age"] == 42

    def test_xlsx_max_rows(self) -> None:
        rows = [[i] for i in range(100)]
        content = _make_xlsx_bytes(["val"], rows)
        result = parse_file(content, "data.xlsx", max_rows=10)
        assert result.row_count == 10

    def test_xlsx_specific_sheet(self) -> None:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Metadata"
        ws1.append(["info"])
        ws1.append(["v1"])
        ws2 = wb.create_sheet("Data")
        ws2.append(["nom", "age"])
        ws2.append(["Dupont", 42])
        buf = io.BytesIO()
        wb.save(buf)
        result = parse_file(buf.getvalue(), "data.xlsx", sheet_name="Data")
        assert result.row_count == 1
        assert result.rows[0]["nom"] == "Dupont"

    def test_xlsx_missing_sheet(self) -> None:
        content = _make_xlsx_bytes(["col"], [[1]])
        with pytest.raises(FileParseError, match="not found"):
            parse_file(content, "data.xlsx", sheet_name="NonExistent")

    def test_xlsx_invalid_magic(self) -> None:
        with pytest.raises(FileParseError, match="XLSX signature"):
            parse_file(b"not an xlsx file content", "data.xlsx")

    def test_xlsx_duplicate_columns(self) -> None:
        content = _make_xlsx_bytes(["nom", "nom"], [[1, 2]])
        with pytest.raises(FileParseError, match="Duplicate"):
            parse_file(content, "data.xlsx")

    def test_xlsx_empty_rows_skipped(self) -> None:
        content = _make_xlsx_bytes(["val"], [[1], [None], [3]])
        result = parse_file(content, "data.xlsx")
        assert result.row_count == 2

    def test_xlsx_none_header_gets_unnamed(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["nom", None, "age"])
        ws.append(["Dupont", "x", 42])
        buf = io.BytesIO()
        wb.save(buf)
        result = parse_file(buf.getvalue(), "data.xlsx")
        assert "_unnamed_1" in result.source_columns

    def test_xlsx_corrupted_after_magic(self) -> None:
        content = b"PK\x03\x04" + b"\x00" * 100
        with pytest.raises(FileParseError, match="Failed to open XLSX"):
            parse_file(content, "data.xlsx")

    def test_xlsx_no_active_sheet(self) -> None:
        wb = Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        with patch("app.services.file_parser.load_workbook") as mock_lwb:
            mock_wb = MagicMock()
            mock_wb.active = None
            mock_wb.sheetnames = ["Sheet1"]
            mock_lwb.return_value = mock_wb
            with pytest.raises(FileParseError, match="no active sheet"):
                parse_file(content, "data.xlsx")
            mock_wb.close.assert_called_once()


# ── parse_file — Format hint ─────────────────────────────


class TestParseFileFormatHint:
    def test_format_hint_csv(self) -> None:
        content = _make_csv_bytes("a;b\n1;2\n")
        result = parse_file(content, "noext", format_hint="csv")
        assert result.detected_format == "csv"
        assert result.row_count == 1

    def test_format_hint_xlsx(self) -> None:
        content = _make_xlsx_bytes(["a"], [[1]])
        result = parse_file(content, "noext", format_hint="xlsx")
        assert result.detected_format == "xlsx"

    def test_format_hint_lucca(self) -> None:
        content = _make_csv_bytes("a;b\n1;2\n")
        result = parse_file(content, "data.txt", format_hint="lucca")
        assert result.detected_format == "lucca"

    def test_format_hint_payfit(self) -> None:
        content = _make_csv_bytes("a;b\n1;2\n")
        result = parse_file(content, "data.txt", format_hint="payfit")
        assert result.detected_format == "payfit"

    def test_unsupported_extension(self) -> None:
        with pytest.raises(FileParseError, match="Unsupported"):
            parse_file(b"data", "file.pdf")

    def test_no_extension_falls_to_csv_path(self) -> None:
        content = _make_csv_bytes("a;b\n1;2\n")
        result = parse_file(content, "noext")
        assert result.detected_format == "csv"


# ── parse_file — edge cases ──────────────────────────────


class TestParseFileEdgeCases:
    def test_no_columns_detected(self) -> None:
        with (
            patch("app.services.file_parser._detect_encoding", return_value="utf-8"),
            patch("app.services.file_parser._decode_content", return_value=""),
            patch("app.services.file_parser._parse_csv", return_value=([], [])),
        ):
            with pytest.raises(FileParseError, match="No columns") as exc_info:
                parse_file(b"something", "data.csv")
            assert exc_info.value.code == "NO_COLUMNS"

    def test_parse_result_is_frozen(self) -> None:
        content = _make_csv_bytes("a;b\n1;2\n")
        result = parse_file(content, "data.csv")
        with pytest.raises(AttributeError):
            result.row_count = 999

    def test_french_date_single_digit(self) -> None:
        content = _make_csv_bytes("date\n1/1/2025\n")
        result = parse_file(content, "data.csv")
        assert result.rows[0]["date"] == datetime.date(2025, 1, 1)


# ── FileParseError ───────────────────────────────────────


class TestFileParseError:
    def test_default_code(self) -> None:
        err = FileParseError("something failed")
        assert err.message == "something failed"
        assert err.code == "FILE_PARSE_ERROR"
        assert str(err) == "something failed"

    def test_custom_code(self) -> None:
        err = FileParseError("empty", code="EMPTY_FILE")
        assert err.code == "EMPTY_FILE"


# ── ParseResult ──────────────────────────────────────────


class TestParseResult:
    def test_frozen_dataclass(self) -> None:
        pr = ParseResult(
            rows=[{"a": 1}],
            source_columns=["a"],
            detected_format="csv",
            detected_encoding="utf-8",
            row_count=1,
        )
        assert pr.row_count == 1
        assert pr.warnings == []
        with pytest.raises(AttributeError):
            pr.rows = []

    def test_defaults(self) -> None:
        pr = ParseResult(
            rows=[],
            source_columns=["a"],
            detected_format="csv",
            detected_encoding="utf-8",
        )
        assert pr.warnings == []
        assert pr.row_count == 0
